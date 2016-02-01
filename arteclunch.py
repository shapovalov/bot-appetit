#!/usr/bin/env python
# -*- coding: utf-8 -*- 

import sys

import openpyxl

MAX_COL = 100
MAX_ROW = 300
HEADER = 3

MAX_LHS_WIDTH = 3
COURSE_COL = 2
FIRST_NAME_COL = 4
WDAY_COL = 2

WDAYS = {u"понедельник" : 1, u"вторник" : 2, u"среда" : 3, u"четверг" : 4, u"пятница" : 5}

def get_pers_columns(ws, name_query):
    pers_columns = {}
    for i in xrange(FIRST_NAME_COL, MAX_COL):
        cell = ws.cell(row = HEADER, column = i).value
        if cell and name_query.lower() in cell.lower():
            pers_columns[i] = cell
    return pers_columns

def enumerate_persons(ws):
    for i in xrange(FIRST_NAME_COL, MAX_COL):
        cell = ws.cell(row = HEADER, column = i).value
        if cell:
            yield (cell.lower(), i)


def split_days(ws):
    split_rows = {}
    last_split = None
    for i in xrange(1, MAX_ROW):
        cell = ws.cell(row = i, column = WDAY_COL).value
        for wd in WDAYS:
            if cell and wd in unicode(cell).lower():
                if last_split is not None:
                    split_rows[last_split[0]] = (last_split[1], i)
                last_split = (WDAYS[wd], i+1)
                break
                    
    split_rows[last_split[0]] = (last_split[1], MAX_ROW)

    if len(split_rows) > len(WDAYS):
        print split_rows
        raise ValueError("Repeated weekdays detected")
        
    return split_rows

def get_order(ws, split_rows, day, col):
    order = []
    for row in xrange(split_rows[day][0], split_rows[day][1]):
        cell = ws.cell(row = row, column = col).value
        if cell:
            course = ws.cell(row = row, column = COURSE_COL).value
            mass = ws.cell(row = row+1, column = 1).value
            num_ord = ws.cell(row = row+1, column = COURSE_COL+1).value
            desc = u""
            if not mass and not num_ord:
                desc = ws.cell(row = row+1, column = COURSE_COL).value
                if desc is None:
                    desc = u""
                
            order.append((course, desc))
            
    return order

def print_order(order, name, verbose=False):
    print u"Here is the order for Mr/Ms", name+u":"
    
    for item in order:
        print item[0]
        if verbose and item[1]:
            print "*", item[1]


if __name__=="__main__":
    if len(sys.argv) < 4:
        print "Incorrect format: pass at least filename, person name and weakday number."
        sys.exit()

    fname = sys.argv[1]
    name = sys.argv[2].decode("utf-8")
    day = int(sys.argv[3])
    verbose = len(sys.argv) > 4 and sys.argv[4] == "-v"

    if day not in xrange(1, 6):
        print "Only working days (1--5) are currently supported."
        sys.exit()

    wb = openpyxl.load_workbook(filename = fname, read_only=True)
    ws = wb.active

    splits = split_days(ws)
    if day not in splits:
        print "Menu not found for the specified day."
        sys.exit()
    columns = get_pers_columns(ws, name)

    print "Found", len(columns), "person(s) matching the string."
    for pers in columns:
        order = get_order(ws, splits, day, pers)
        print ""
        print_order(order, columns[pers], verbose=verbose)
